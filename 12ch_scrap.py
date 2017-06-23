import sys, pdb
sys.path.append('/usr/share/doc')
sys.path.append("/usr/lib/python3/dist-packages")
sys.path.append("/usr/local/lib/python3.4/dist-packages")
sys.path.append("/usr/local/lib/python2.7/dist-packages")
import numpy as np
import pandas as pd
import xlrd, xlwt
import xlsxwriter
import openpyxl as oxl
import string
from utils import timeme
path = 'data/'

def gen_wbs():
    wb = xlwt.Workbook()
    print(wb)
    print(wb.add_sheet('first_sheet', cell_overwrite_ok=True))
    print(wb.get_active_sheet())
    ws_1 = wb.get_sheet(0)
    print(ws_1)
    ws_2 = wb.add_sheet('second_sheet')
    data = np.arange(1, 65).reshape((8, 8))
    print(data)
    ws_1.write(0, 0, 100)
    wb.save(path + 'workbook.xls')

def gen_wbs2():
    data = np.arange(1, 65).reshape((8, 8))
    wb = xlsxwriter.Workbook(path + 'workbook.xlsx')
    ws_1 = wb.add_worksheet('first_sheet')
    ws_2 = wb.add_worksheet('second_sheet')
    for c in range(data.shape[0]):
        for r in range(data.shape[1]):
            ws_1.write(r, c, data[c, r])
            ws_2.write(r, c, data[r, c])
    wb.close()
    wb = xlsxwriter.Workbook(path + 'chart.xlsx')
    ws = wb.add_worksheet()

    # write cumsum of random values in first column
    values = np.random.standard_normal(15).cumsum()
    ws.write_column('A1', values)

    # create a new chart object
    chart = wb.add_chart({'type': 'line'})

    # add a series to the chart
    chart.add_series({'values': '=Sheet1!$A$1:$A$15',
                      'marker': {'type': 'diamond'},})
    # series with markers (here: diamond)

    # insert the chart
    ws.insert_chart('C1', chart)
    wb.close()

def reading_wbs():
    book = xlrd.open_workbook(path + 'workbook.xlsx')
    print(book)
    print(book.sheet_names())
    sheet_1 = book.sheet_by_name('first_sheet')
    sheet_2 = book.sheet_by_index(1)
    print(sheet_1)
    print(sheet_2.name)
    print(sheet_1.ncols, sheet_1.nrows)
    cl = sheet_1.cell(0, 0)
    print(cl.value)
    print(cl.ctype)
    print(sheet_2.row(3))
    print(sheet_2.col(3))
    print(sheet_1.col_values(3, start_rowx=3, end_rowx=7))
    print(sheet_1.row_values(3, start_colx=3, end_colx=7))
    
    for c in range(sheet_1.ncols):
        for r in range(sheet_1.nrows):
            print ('%i' % sheet_1.cell(r, c).value, end=' ')
            
def using_openpyxl():
    data = np.arange(1, 65).reshape((8, 8))
    wb = oxl.Workbook()
    ws = wb.create_sheet(index=0, title='oxl_sheet')
    for c in range(data.shape[0]):
        for r in range(data.shape[1]):
            ws.cell(row=r+1, column=c+1).value = data[c, r]
            # creates a Cell object and assigns a value
    wb.save(path + 'oxl_book.xlsx')
    wb = oxl.load_workbook(path + 'oxl_book.xlsx')
    ws = wb.get_active_sheet()
    cell = ws['B4']
    print(cell.column)
    print(cell.row)
    print(cell.value)
    print(ws['B1':'B4'])

    for cell in ws['B1':'B4']:
        print(cell[0].value)

    for row in ws['B1':'B4']:
        for cell in row:
            print(cell.value, end=' ')
        print

def pandas_rw():
    data = np.arange(1, 65).reshape((8, 8))
    df_1 = pd.read_excel(path + 'workbook.xlsx',
                     'first_sheet', header=None)
    df_2 = pd.read_excel(path + 'workbook.xlsx',
                     'second_sheet', header=None)
    columns = []
    for c in range(data.shape[0]):
        columns.append(string.ascii_uppercase[c])
    print(columns)
    df_1.columns = columns
    df_2.columns = columns
    print(df_1)
    print(df_2)
    df_1.to_excel(path + 'new_book_1.xlsx', 'my_sheet')
    wbn = xlrd.open_workbook(path + 'new_book_1.xlsx')
    print(wbn.sheet_names())
    wbw = pd.ExcelWriter(path + 'new_book_2.xlsx')
    df_1.to_excel(wbw, 'first_sheet')
    df_2.to_excel(wbw, 'second_sheet')
    wbw.save()
    wbn = xlrd.open_workbook(path + 'new_book_2.xlsx')
    print(wbn.sheet_names())
    data = np.random.rand(20, 10000)
    print(data.nbytes)
    df = pd.DataFrame(data)
    timeme(df.to_excel)(path + 'data.xlsx', 'data_sheet')
    timeme(np.save)(path + 'data', data)
    df = timeme(pd.read_excel)(path + 'data.xlsx', 'data_sheet')
    data = timeme(np.load)(path + 'data.npy')
    data, df = 0.0, 0.0

if __name__ == '__main__':
    # gen_wbs()
    # gen_wbs2()
    # reading_wbs()
    # using_openpyxl()
    pandas_rw()